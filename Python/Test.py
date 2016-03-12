'''
Created on Feb 23, 2016

@author: Devin
'''

from openpyxl import load_workbook
import numpy as np

wb = load_workbook(filename = 'Areas.xlsx')

sheet = wb.get_sheet_by_name('Sarah')

type(sheet)

d = [sheet.cell(row = i, column = 2).value for i in range(1, 7)]


def close_enough(nextVal, currentVal, error = 1e-7):
    diff = nextVal - currentVal
    diff = abs(diff)
    maxdiff = max(diff)
    if maxdiff < error:
        return True
    else:
        return False

def calc(x):
    if(x == ">"):
        return 3
    elif(x == ">>"):
        return 9
    elif(x == "<"):
        return 1./3
    elif(x == "<<"):
        return 1./9
    elif(x == "E"):
        return 1
    else:
        return("Error, unknown value "+ x)

def fullCalc(x):
    return Largest_eigen(outMat(x))


def outMat(x):
    xsheet = wb.get_sheet_by_name(x)
    type(xsheet)
    f = [xsheet.cell(row = i, column = 2).value for i in range(1, 7)]
    outputArray = np.array([calc(j) for j in f])
    ab = outputArray[0]
    bc = outputArray[1]
    cd = outputArray[2]
    ac = outputArray[3]
    bd = outputArray[4]
    ad = outputArray[5]
    ba = 1./outputArray[0]
    cb = 1./outputArray[1]
    dc = 1./outputArray[2]
    ca = 1./outputArray[3]
    db = 1./outputArray[4]
    da = 1./outputArray[5]

    return np.array([[1, ab, ac, ad],
                     [ba, 1 , bc,bd],
                     [ca,  cb, 1, cd],
                     [da, db, dc, 1]])

def Largest_eigen(x, error = 1e-7):
    size = x.shape[0]
    currentVal = np.array([1./size for i in range(size)])
    while True:
        nextVal = np.matmul(x, currentVal)
        nextVal = nextVal / sum(nextVal)
        #print "Working Still"
        if close_enough(nextVal, currentVal, error = 1e-7):
            return nextVal
        currentVal = nextVal
    return currentVal

print fullCalc("Drew")


