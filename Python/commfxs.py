'''
Created on Feb 23, 2016

@author: Devin
'''
from plotly.offline import plot
from plotly.offline import download_plotlyjs, init_notebook_mode, iplot
import plotly.graph_objs as go
from openpyxl import load_workbook
import numpy as np

#wb = load_workbook(filename = 'Areas.xlsx')

#sheet = wb.get_sheet_by_name('Sarah')
#type(sheet)

#d = [sheet.cell(row = i, column = 2).value for i in range(1, 7)]


def close_enough(nextVal, currentVal, error = 1e-7):
    diff = nextVal - currentVal
    diff = abs(diff)
    maxdiff = max(diff)
    if maxdiff < error:
        return True
    else:
        return False

def calc(x, better = 3.0, muchbetter = 9.0):
    if(x == ">"):
        return 1./better
    elif(x == ">>"):
        return 1./muchbetter
    elif(x == "<"):
        return better
    elif(x == "<<"):
        return muchbetter
    elif(x == "E"):
        return 1
    else:
        return("Error, unknown value "+ x)

def single_stats(fname, sheetname, bars = False, better = 3,  muchbetter = 9):
    if bars:
        eigen = single_stats(fname, sheetname, bars = False, better = better, muchbetter = muchbetter)
        altnames = get_altnames("Areas.xlsx")
        data = go.Bar(x=altnames, y=eigen)
        layout = go.Layout(title = sheetname+"'s Priorities")
        return iplot(go.Figure(data = go.Data([data]), layout = layout))

    else:
        return Largest_eigen(getMatrix(fname, sheetname, better, muchbetter))

def group_stats(xlsxFname, listOfSheetnames, bars = False):
    if bars:
        barVals = group_stats(xlsxFname, listOfSheetnames, bars = False)
        # Do nifty stuff like in tutorial to create plotly chart
        return "I should do bars, but I haven't yet"
    else:
        listOfMatrices = []
        for sheetName in listOfSheetnames:
            theMatrix = getMatrix(xlsxFname, sheetName)
            listOfMatrices.append(theMatrix)
        geomAvg = geometric_avg(listOfMatrices)
        return Largest_eigen(geomAvg)

def Largest_eigen(x, error = 1e-7):
    size = x.shape[0]
    currentVal = np.array([1./size for i in range(size)])
    while True:
        nextVal = np.matmul(x, currentVal)
        nextVal = nextVal / sum(nextVal)
        #print "Working Still"
        if close_enough(nextVal, currentVal, error = 1e-7):
            return nextVal
        currentVal = nextVal
    return currentVal

def get_altnames(xlsxfile):
    wb = load_workbook(filename = xlsxfile)
    firstsheet = wb.get_sheet_names()[0]
    sheet = wb.get_sheet_by_name(firstsheet)
    nrows = sheet.max_row
    firstcol = [sheet.cell(row = i, column = 1).value for i in range(1, nrows)]
    thirdcol = [sheet.cell(row = i, column = 3).value for i in range(1, nrows)]

    rval = list()
    for row in range(0, len(firstcol)):
        if thirdcol[row] != None:
            if (firstcol[row] != None) and (firstcol[row] not in rval):
                rval.append(firstcol[row])
            if (thirdcol[row] not in rval):
                rval.append(thirdcol[row])
    return rval
    
def listtest(x):   
    list1 = list([1, 2, 5, 6, 9])
    if x not in list1:
        return "number not found"
    else:
        return "number was found"


#print get_altnames('Areas.xlsx')
#print listtest(6)

def getMatrix(fname, sheetName, better = 3, muchbetter = 9):
    #First we need to know the alternatives in this sheet
    altnames = get_altnames(fname)
    #The number of alternatives is the size of the return matrix
    matSize = len(altnames)
    #The return matrix starts off with 1's on the diagonal, and zeroes elsewhere
    returnMatrix = np.identity(matSize)
    #Load up the excel spreadsheet and get the actual sheet
    wb = load_workbook(filename = fname)
    xsheet = wb.get_sheet_by_name(sheetName)
    #Loop over the rows in this sheet
    for row in range(xsheet.max_row):
        #Get the first column val, which is the row of the comparison
        val1 = xsheet.cell(row = row+1, column = 1).value
        #Get the second column val, which is the value of the comparison
        val2 = xsheet.cell(row = row+1, column = 2).value
        #Get the third column value, which is the COLUMN of the comparison
        val3 = xsheet.cell(row = row+1, column = 3).value
        #If third column value == None, that means this wasn't a pairwise comparison entry
        #Maybe it was a demographic bit (FavColor)
        if (val3 != None):
            #Get the index of the alt name in the list of altnames
            #This tells us where to place the vote in the resulting matrix
            rowIndex = altnames.index(val1)
            colIndex = altnames.index(val3)
            #Get the numeric value of the vote from the stringy vote
            voteValue = calc(val2, better, muchbetter)
            #Lastly place the vote in, and it's reciprocal on the other side
            #of the diagonal
            returnMatrix[rowIndex,colIndex] = voteValue
            returnMatrix[colIndex, rowIndex] = 1./voteValue
#            returnMatrix[index3, index1] = 
    return returnMatrix


#def getBars(stats, barnames, title, xlsxfile, sheetname):
    #altnames = get_altnames(xlsxfile)
    #data = go.Bar(x= altnames, y= single_stats(xlsxfile, sheetname))
    ##layout = go.Layout(title = sheetname+ "'s Priorities")
    #iplot(go.Figure(data = go.Data([data]), layout = layout))
    
def geometric_avg(listOfMats):
    #Create rval
    rval = np.zeros_like(listOfMats[0])
    nrows = rval.shape[0]
    ncols = rval.shape[1]
    for row in range(nrows):
        for col in range(ncols):
            val = 1
            count = 0
            for mat in listOfMats:
                if mat[row, col]!=0:
                    val *= mat[row,col]
                    count+=1
            if count != 0:
                val = pow(val, 1.0/count)
            #Finally have the value, put it in
            rval[row,col] = val
    return rval
print group_stats('Areas.xlsx', ['Sarah', 'Drew'])       
#print getMatrix('Areas.xlsx', 'Sarah')
