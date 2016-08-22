'''
Created on Aug 20, 2016

@author: wjadams
'''
from openpyxl import load_workbook
import numpy as np
from all_user_votes import PairwiseAllUsers, inv_vote

################################################################
####Begin code to parse Will's Excel File Format  ##############
################################################################

def get_altnames_will_excel(xlsxfile):
    wb = load_workbook(filename = xlsxfile)
    firstsheet = wb.get_sheet_names()[0]
    sheet = wb.get_sheet_by_name(firstsheet)
    nrows = sheet.max_row
    firstcol = [sheet.cell(row = i, column = 1).value for i in range(1, nrows)]
    thirdcol = [sheet.cell(row = i, column = 3).value for i in range(1, nrows)]

    rval = list()
    for row in range(0, len(firstcol)):
        if thirdcol[row] != None:
            if (firstcol[row] != None) and (firstcol[row] not in rval):
                rval.append(firstcol[row])
            if (thirdcol[row] not in rval):
                rval.append(thirdcol[row])
    return rval

def get_usernames_will_excel(xlsxFile):
    wb = load_workbook(filename = xlsxFile)
    return wb.get_sheet_names()
    
def calc(x):
    if(x == ">"):
        return -1.5
    elif(x == ">>"):
        return -2.5
    elif(x == "<"):
        return -1
    elif(x == "<<"):
        return -2
    elif(x == "E"):
        return 1
    else:
        return("Error, unknown value "+ x)
    
def get_matrix_from_will(fname, sheetName):
    #First we need to know the alternatives in this sheet
    altnames = get_altnames_will_excel(fname)
    #The number of alternatives is the size of the return matrix
    matSize = len(altnames)
    #The return matrix starts off with 1's on the diagonal, and zeroes elsewhere
    returnMatrix = np.identity(matSize)
    #Load up the excel spreadsheet and get the actual sheet
    wb = load_workbook(filename = fname)
    xsheet = wb.get_sheet_by_name(sheetName)
    #Loop over the rows in this sheet
    for row in range(xsheet.max_row):
        #Get the first column val, which is the row of the comparison
        val1 = xsheet.cell(row = row+1, column = 1).value
        #Get the second column val, which is the value of the comparison
        val2 = xsheet.cell(row = row+1, column = 2).value
        #Get the third column value, which is the COLUMN of the comparison
        val3 = xsheet.cell(row = row+1, column = 3).value
        #If third column value == None, that means this wasn't a pairwise comparison entry
        #Maybe it was a demographic bit (FavColor)
        if (val3 != None):
            #Get the index of the alt name in the list of altnames
            #This tells us where to place the vote in the resulting matrix
            rowIndex = altnames.index(val1)
            colIndex = altnames.index(val3)
            #Get the numeric value of the vote from the stringy vote
            voteValue = calc(val2)
            #Lastly place the vote in, and it's reciprocal on the other side
            #of the diagonal
            returnMatrix[rowIndex,colIndex] = voteValue
            returnMatrix[colIndex, rowIndex] = inv_vote(voteValue)
#            returnMatrix[index3, index1] = 
    return returnMatrix

def from_will_excel(xlsxFile):
    allvotes = PairwiseAllUsers()
    allvotes.clear()
    allvotes.alt_names = get_altnames_will_excel(xlsxFile)
    for user in get_usernames_will_excel(xlsxFile):
        allvotes.add_user(user, get_matrix_from_will(xlsxFile, user))
    return(allvotes)
################################################################
#### End code to parse Will's Excel File Format ################
################################################################



votes2 = from_will_excel("Areas.xlsx")
print(votes2.get_matrix_raw("Sarah"))
print(votes2.get_matrix("Sarah"))
print(votes2.single_stats("Sarah"))
print(votes2.single_stats("Sarah", doppelganger=True))
print(votes2.get_matrix("Sarah", sym_vote_values = [2, 3]))
print(votes2.single_stats("Sarah", sym_vote_values = [2, 3]))
