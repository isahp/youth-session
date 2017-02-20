from openpyxl import load_workbook
import numpy as np
from all_user_votes import PairwiseAllUsers, inv_vote
import sys

def get_num_data(aval):
    if sys.version_info < (3,0):
        NumberTypes = (int, float, complex, long)
        StringTypes = (str, unicode)
    else:
        NumberTypes = (int, float, complex)
        StringTypes = (str)
    if isinstance(aval, NumberTypes):
        return aval
    elif isinstance(aval, StringTypes):
        if aval.startswith("="):
            aval=aval[1:]
        #print("Found string")
        if "/" in aval:
            num_denom = aval.split("/")
            num = float(num_denom[0])
            denom = float(num_denom[1])
            #print(num_denom)
            rval = num/denom
        else:
            #Try to just parse it
            try:
                rval = float(aval)
            except:
                #Didn't parse, so we'll call it 0
                rval = 0
        return rval
    else:
        print("Unknown")
        return 0.0

def get_nish_matrix(sheet, row, compare_cols, nalts):
    rval = np.identity(nalts)
    for compare_col in compare_cols:
        sheet_col = compare_col[0]
        data = sheet.cell(row=row, column=sheet_col).value
        num_data = get_num_data(data)
        #print(data)
        #print(num_data)
        first_alt = compare_col[1]
        second_alt = compare_col[2]
        rval[first_alt, second_alt] = num_data + 0.0
        if num_data != 0.0:
            rval[second_alt, first_alt] = 1.0/num_data
        else:
            rval[second_alt, first_alt] = 0.0
    return(rval)


def from_nish_excel(xlsxFile, sheetName = "Weight"):
    wb = load_workbook(xlsxFile, read_only=True)
    sheet = wb.get_sheet_by_name(sheetName)
    ncol = sheet.max_column
    nrow = sheet.max_row
    firstcol = [sheet.cell(row=i, column=1).value for i in range(1, nrow)]
    #The row holding headers is the one with first non-None value in 1st column
    for i in range(1, nrow):
        if (firstcol[i] != None):
            row_offset = i+1
            break
    #Now that we know the row_offset, we can figure out which columns have Name1 v Name2
    #format?
    header_row = [sheet.cell(row=row_offset, column=i).value for i in range(1,ncol)]
    alt_names = []
    compare_cols = []
    for i in range(len(header_row)):
        thestr = header_row[i].lower().strip()
        info = thestr.split(" v ")
        if len(info) == 2:
            #We have found a comparison header
            if not info[0] in alt_names:
                alt_names.append(info[0])
            if not info[1] in alt_names: 
                alt_names.append(info[1])
            #We have a comparison in this column
            #The spreadsheet column indexing starts at 1, so the actual column is 1 more than i
            col = i + 1
            #First alt is the index of alt1
            first_alt = alt_names.index(info[0])
            second_alt = alt_names.index(info[1])
            data = [col, first_alt, second_alt]
            compare_cols.append(data)
    #Now setup the return value
    rval = PairwiseAllUsers()
    nalts = len(alt_names)
    print(alt_names)
    print(compare_cols)
    #Create users
    user_names = []
    user_rows= []
    for row in range(row_offset+1, nrow):
        uid = sheet.cell(row=row, column=2).value
        if uid in user_names:
            raise NameError("Duplicate user "+uid+" I give up")
        user_names.append(uid)
        user_rows.append(row)
        pcmp = get_nish_matrix(sheet, row, compare_cols, nalts)
        rval.add_user(uid, pcmp)
    print(user_names)
    print(user_rows)
    #Get usernames
    rval.alt_names = alt_names
    #offset = find_offset()
    #endrow = find_end_row()
    #im going to start with psuedo code
    return(rval)

